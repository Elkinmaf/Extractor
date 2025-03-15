#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
excel_manager.py - Gestión de archivos Excel para SAP Issues Extractor
---
Este módulo proporciona una clase para manejar los archivos Excel usados para
almacenar y dar seguimiento a los issues extraídos de SAP.
"""

import os
import logging
from datetime import datetime
from tkinter import filedialog, messagebox

# Importaciones opcionales con gestión de errores
try:
    import pandas as pd
    import numpy as np
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False
    pd = None
    np = None

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    OPENPYXL_AVAILABLE = False
except ImportError:
    OPENPYXL_AVAILABLE = False

# Configurar logger
logger = logging.getLogger(__name__)

class ExcelManager:
    """
    Clase dedicada al manejo de archivos Excel de seguimiento de issues.
    
    Esta clase proporciona métodos para seleccionar o crear archivos Excel,
    actualizar datos, aplicar formato y gestionar el tracking de issues.
    """
    
    def __init__(self, file_path=None):
        """
        Inicializa el administrador de Excel
        
        Args:
            file_path (str, optional): Ruta al archivo Excel. Si es None, se pedirá al usuario.
        """
        self.file_path = file_path
        
        # Verificar disponibilidad de pandas
        if not PANDAS_AVAILABLE:
            logger.warning("Pandas no está instalado. La funcionalidad de Excel será limitada.")
        
        # Verificar disponibilidad de openpyxl
        if not OPENPYXL_AVAILABLE:
            logger.warning("Openpyxl no está instalado. El formato de Excel será limitado.")
            
        # Columnas predefinidas para el archivo Excel
        self.default_columns = [
            "Title",
            "Type",
            "Priority",
            "Status",
            "Deadline",
            "Due Date",
            "Created By",
            "Created On",
            "Last Updated",
            "Comments",
        ]
        
    def select_file(self):
        """
        Permite al usuario elegir un archivo Excel existente o crear uno nuevo
        
        Returns:
            str: Ruta al archivo Excel seleccionado o creado
        """
        # Preguntar si desea usar un archivo existente o crear uno nuevo
        choice = messagebox.askquestion(
            "Archivo Excel",
            "¿Desea usar un archivo Excel existente?\n\n"
            + "Seleccione 'Sí' para elegir un archivo existente.\n"
            + "Seleccione 'No' para crear un nuevo archivo.",
            icon='question'
        )

        if choice == "yes":
            # Permitir al usuario seleccionar un archivo Excel existente
            file_path = filedialog.askopenfilename(
                title="Seleccione el archivo Excel de seguimiento",
                filetypes=[("Archivos Excel", "*.xlsx"), ("Todos los archivos", "*.*")],
            )

            if not file_path:  # El usuario canceló la selección
                logger.info("Usuario canceló la selección de archivo. Se creará uno nuevo.")
                self._create_new_file()
            else:
                self.file_path = file_path
                logger.info(f"Archivo Excel seleccionado: {file_path}")
        else:
            self._create_new_file()
            
        return self.file_path
    
    def _create_new_file(self):
        """
        Crea un nuevo archivo Excel para seguimiento de issues
        
        Returns:
            str: Ruta al archivo Excel creado
        """
        # Verificar que pandas esté disponible
        if not PANDAS_AVAILABLE:
            messagebox.showerror(
                "Error", 
                "No se puede crear un nuevo archivo Excel porque la biblioteca pandas no está instalada.\n"
                "Por favor, instale pandas con: pip install pandas openpyxl"
            )
            return None
            
        # Crear un nombre de archivo por defecto con fecha y hora
        default_filename = f"Seguimiento_Issues_SAP_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        # Permitir al usuario guardar con un nombre específico
        file_path = filedialog.asksaveasfilename(
            title="Guardar nuevo archivo Excel",
            defaultextension=".xlsx",
            initialfile=default_filename,
            filetypes=[("Archivos Excel", "*.xlsx")],
        )

        if not file_path:  # Si cancela, usar el nombre por defecto
            # Determinar la ruta de Documentos
            documents_path = os.path.join(os.path.expanduser("~"), "Documents")
            if not os.path.exists(documents_path):
                documents_path = os.path.join(os.path.expanduser("~"), "Documentos")
            if not os.path.exists(documents_path):
                documents_path = os.path.expanduser("~")
                
            file_path = os.path.join(documents_path, default_filename)
            logger.info(f"Se usará el nombre por defecto: {file_path}")
            
            # Informar al usuario
            messagebox.showinfo(
                "Archivo Nuevo", 
                f"Se creará un nuevo archivo Excel con el nombre predeterminado:\n{file_path}"
            )

        # Crear un archivo Excel vacío con las columnas necesarias
        success = self._create_excel_template(file_path)
        
        if success:
            logger.info(f"Creado nuevo archivo Excel: {file_path}")
            self.file_path = file_path
            return file_path
        else:
            logger.error(f"No se pudo crear el archivo Excel: {file_path}")
            return None
        
    def _create_excel_template(self, file_path):
        """
        Crea la estructura del archivo Excel con las columnas necesarias
        
        Args:
            file_path (str): Ruta donde crear el archivo Excel
            
        Returns:
            bool: True si se creó correctamente, False en caso contrario
        """
        try:
            # Crear un DataFrame vacío con las columnas necesarias
            df = pd.DataFrame(columns=self.default_columns)

            # Guardar el DataFrame vacío como un archivo Excel
            df.to_excel(file_path, index=False, engine='openpyxl')
            
            # Si openpyxl está disponible, aplicar formato básico
            if OPENPYXL_AVAILABLE:
                self._apply_excel_formatting(file_path)
                
            logger.info(f"Archivo Excel creado exitosamente: {file_path}")
            return True
        except Exception as e:
            logger.error(f"Error al crear nuevo archivo Excel: {e}")
            messagebox.showerror("Error", f"No se pudo crear el archivo Excel: {e}")
            return False
            
    def update_with_issues(self, issues_data):
        """
        Actualiza el archivo Excel con los datos extraídos
        
        Args:
            issues_data (list): Lista de diccionarios con datos de issues
            
        Returns:
            tuple: (success, new_items, updated_items)
        """
        if not PANDAS_AVAILABLE:
            logger.error("No se puede actualizar Excel: pandas no está instalado")
            messagebox.showerror(
                "Error", 
                "No se puede actualizar el archivo Excel porque la biblioteca pandas no está instalada.\n"
                "Por favor, instale pandas con: pip install pandas openpyxl"
            )
            return False, 0, 0
            
        if not self.file_path:
            logger.error("No hay ruta de archivo Excel especificada")
            messagebox.showerror("Error", "No se ha seleccionado un archivo Excel para guardar los datos")
            return False, 0, 0
            
        if not issues_data:
            logger.warning("No hay datos para actualizar en Excel")
            messagebox.showwarning("Sin datos", "No hay datos de issues para actualizar en el Excel")
            return False, 0, 0
            
        try:
            logger.info(f"Actualizando archivo Excel: {self.file_path}")
            
            # Cargar el archivo existente o crear estructura si no existe
            if os.path.exists(self.file_path):
                try:
                    existing_df = pd.read_excel(self.file_path, engine='openpyxl')
                    logger.info(f"Archivo Excel existente cargado con {len(existing_df)} registros")
                except Exception as read_e:
                    logger.warning(f"Error al leer Excel: {read_e}. Creando estructura nueva.")
                    existing_df = pd.DataFrame(columns=self.default_columns)
            else:
                existing_df = pd.DataFrame(columns=self.default_columns)
                logger.info("Creando nueva estructura de Excel")

            # Convertir datos de issues a DataFrame
            new_df = pd.DataFrame(issues_data)
            
            # Contadores para estadísticas
            new_items = 0
            updated_items = 0
            
            # Hacer una copia del DataFrame existente para no modificarlo mientras iteramos
            updated_df = existing_df.copy()
            
            # Optimización: crear índice para búsquedas rápidas si hay muchos registros
            title_index = {}
            if len(existing_df) > 0 and "Title" in existing_df.columns:
                for idx, title in enumerate(existing_df["Title"]):
                    if title and not pd.isna(title) and title not in title_index:
                        title_index[title] = idx
            
            # Procesar cada issue nuevo
            for _, new_row in new_df.iterrows():
                title = new_row.get("Title", "")
                if not title or pd.isna(title):
                    continue  # Saltar filas sin título
                    
                title_exists = title in title_index
                
                if not title_exists:
                    # Agregar fecha de última actualización para elementos nuevos
                    new_row_dict = new_row.to_dict()
                    new_row_dict["Last Updated"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    
                    # Asegurar que todos los campos esperados existan
                    for column in self.default_columns:
                        if column not in new_row_dict:
                            new_row_dict[column] = ""
                            
                    new_row_df = pd.DataFrame([new_row_dict])
                    updated_df = pd.concat([updated_df, new_row_df], ignore_index=True)
                    new_items += 1
                    logger.info(f"Nuevo issue añadido: '{title}'")
                else:
                    # Obtener índice del elemento existente
                    idx = title_index[title]
                    
                    # Verificar cambios en el estado y otras columnas
                    updated = False
                    
                    for column in ["Status", "Priority", "Type", "Due Date", "Deadline", "Created By", "Created On"]:
                        if column in new_row and column in existing_df.columns:
                            old_value = existing_df.iloc[idx][column] if not pd.isna(existing_df.iloc[idx][column]) else ""
                            new_value = new_row[column] if column in new_row and not pd.isna(new_row[column]) else ""
                            
                            if str(old_value) != str(new_value):
                                mask = updated_df["Title"] == title
                                updated_df.loc[mask, column] = new_value
                                updated_df.loc[mask, "Last Updated"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                                updated = True
                                logger.info(f"Actualizado {column} de '{title}': '{old_value}' → '{new_value}'")
                    
                    if updated:
                        updated_items += 1
            
            # Guardar el DataFrame actualizado
            updated_df.to_excel(self.file_path, index=False, engine='openpyxl')
            
            # Aplicar formato al Excel
            if OPENPYXL_AVAILABLE:
                self._apply_excel_formatting()
            
            logger.info(f"Excel actualizado: {new_items} nuevos, {updated_items} actualizados")
            return True, new_items, updated_items
            
        except Exception as e:
            logger.error(f"Error al actualizar el archivo Excel: {e}")
            messagebox.showerror("Error", f"Error al actualizar el archivo Excel: {e}")
            return False, 0, 0
            
    def _apply_excel_formatting(self, file_path=None):
        """
        Aplica formato estético al archivo Excel
        
        Args:
            file_path (str, optional): Ruta al archivo Excel. Si es None, usa self.file_path
            
        Returns:
            bool: True si se aplicó el formato correctamente, False en caso contrario
        """
        if not OPENPYXL_AVAILABLE:
            logger.warning("Openpyxl no está disponible. No se puede aplicar formato.")
            return False
            
        if file_path is None:
            file_path = self.file_path
        
        try:
            wb = load_workbook(file_path)
            ws = wb.active
            
            # Formato para encabezados
            header_fill = PatternFill(
                start_color="1F4E78", end_color="1F4E78", fill_type="solid"
            )
            header_font = Font(bold=True, color="FFFFFF")
            header_alignment = Alignment(horizontal="center", vertical="center")

            thin_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

            # Aplicar formato a encabezados
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=1, column=col)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = thin_border

            # Aplicar formato a celdas de datos
            for row in range(2, ws.max_row + 1):
                for col in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row, column=col)
                    cell.border = thin_border

                    # Colorear por estado
                    if col == 4:  # Columna Status
                        status = str(cell.value).upper() if cell.value else ""

                        if "DONE" in status:
                            cell.fill = PatternFill(
                                start_color="CCFFCC",
                                end_color="CCFFCC",
                                fill_type="solid",
                            )
                        elif "OPEN" in status:
                            cell.fill = PatternFill(
                                start_color="FFCCCC",
                                end_color="FFCCCC",
                                fill_type="solid",
                            )
                        elif "READY" in status:
                            cell.fill = PatternFill(
                                start_color="FFFFCC",
                                end_color="FFFFCC",
                                fill_type="solid",
                            )
                        elif "IN PROGRESS" in status:
                            cell.fill = PatternFill(
                                start_color="FFE6CC",
                                end_color="FFE6CC",
                                fill_type="solid",
                            )

            # Ajustar ancho de columnas
            for col in range(1, ws.max_column + 1):
                max_length = 0
                for row in range(1, ws.max_row + 1):
                    cell_value = ws.cell(row=row, column=col).value
                    if cell_value:
                        max_length = max(max_length, len(str(cell_value)))

                adjusted_width = max(10, min(50, max_length + 2))
                ws.column_dimensions[
                    ws.cell(row=1, column=col).column_letter
                ].width = adjusted_width

            wb.save(file_path)
            logger.info("Formato aplicado al archivo Excel correctamente")
            return True
        except Exception as format_e:
            logger.warning(f"No se pudo aplicar formato al Excel: {format_e}")
            return False
            
    def get_file_path(self):
        """
        Obtiene la ruta del archivo Excel actual
        
        Returns:
            str: Ruta al archivo Excel o None si no hay archivo seleccionado
        """
        return self.file_path
        
    def export_to_csv(self, output_path=None):
        """
        Exporta los datos del Excel a un archivo CSV
        
        Args:
            output_path (str, optional): Ruta de salida para el CSV. Si es None, se deriva del Excel.
            
        Returns:
            bool: True si la exportación fue exitosa, False en caso contrario
        """
        if not PANDAS_AVAILABLE:
            logger.error("No se puede exportar a CSV: pandas no está instalado")
            return False
            
        if not self.file_path:
            logger.error("No hay archivo Excel seleccionado para exportar")
            return False
            
        try:
            # Cargar datos del Excel
            df = pd.read_excel(self.file_path, engine='openpyxl')
            
            # Determinar la ruta de salida
            if not output_path:
                output_path = os.path.splitext(self.file_path)[0] + ".csv"
                
            # Exportar a CSV
            df.to_csv(output_path, index=False, encoding='utf-8-sig')  # Con BOM para Excel
            
            logger.info(f"Datos exportados a CSV: {output_path}")
            return True
        except Exception as e:
            logger.error(f"Error al exportar a CSV: {e}")
            return False
            
    def get_stats(self):
        """
        Obtiene estadísticas básicas del archivo Excel
        
        Returns:
            dict: Diccionario con estadísticas o None si hay error
        """
        if not PANDAS_AVAILABLE or not self.file_path:
            return None
            
        try:
            # Cargar datos del Excel
            df = pd.read_excel(self.file_path, engine='openpyxl')
            
            # Estadísticas básicas
            stats = {
                "total_issues": len(df),
                "by_status": {}
            }
            
            # Contar por estado
            if "Status" in df.columns:
                status_counts = df["Status"].value_counts().to_dict()
                stats["by_status"] = status_counts
                
            # Contar por prioridad
            if "Priority" in df.columns:
                priority_counts = df["Priority"].value_counts().to_dict()
                stats["by_priority"] = priority_counts
                
            # Contar por tipo
            if "Type" in df.columns:
                type_counts = df["Type"].value_counts().to_dict()
                stats["by_type"] = type_counts
                
            # Fechas importantes
            if len(df) > 0:
                stats["last_updated"] = df["Last Updated"].max() if "Last Updated" in df.columns else None
                
            return stats
        except Exception as e:
            logger.error(f"Error al obtener estadísticas: {e}")
            return None

    def open_excel_file(self):
        """
        Abre el archivo Excel con la aplicación predeterminada del sistema
        
        Returns:
            bool: True si se abrió correctamente, False en caso contrario
        """
        if not self.file_path or not os.path.exists(self.file_path):
            logger.error("No hay archivo Excel válido para abrir")
            return False
            
        try:
            import webbrowser
            webbrowser.open(self.file_path)
            logger.info(f"Archivo Excel abierto: {self.file_path}")
            return True
        except Exception as e:
            logger.error(f"Error al abrir archivo Excel: {e}")
            
            # Intento alternativo con os
            try:
                if os.name == 'nt':  # Windows
                    os.startfile(self.file_path)
                elif os.name == 'posix':  # macOS y Linux
                    os.system(f"open {self.file_path}")
                return True
            except Exception as e2:
                logger.error(f"Error alternativo al abrir Excel: {e2}")
                return False
